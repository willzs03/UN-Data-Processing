from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import  Font


wb = load_workbook("MICS_UNICEF_SurveyData.xlsx")
ws = wb.active

# Delete columns B(2), E(5), F(6), G(7)
ws.delete_cols(2)
ws.delete_cols(4, 3)

# Format columns C1 (Start Year) and D1 (End Year) appropriately
for cell in ws['C']:
    currentCol = get_column_letter(cell.column) + str(cell.row)
    nextCol = get_column_letter(cell.column + 1) + str(cell.row)
    if "-" not in cell.value:
        ws[currentCol] = cell.value
        ws[nextCol] = cell.value

    # Cases with a range. e.g. 2011-2015 -> Start Year = 2011, End Year = 2015
    else:
        # Split range into two items. e.g. '2011-2015' -> ['2011', '2015']
        range = cell.value.split('-')
        ws[currentCol] = range[0]
        ws[nextCol] = range[1]

# Update 'Source' column
for cell in ws['D']:
    nextCol = get_column_letter(cell.column + 1) + str(cell.row)
    ws[nextCol] = 'MICS_UNICEF'

ws.move_range("A1:A400", cols=5)
ws.move_range("B1:B400", cols=-1)
ws.move_range("F1:F400", cols=-4)

# Rename column titles for consistency
ws['A1'] = 'nation'
ws['B1'] = 'title'
ws['C1'] = 'year_start'
ws['D1'] = 'year_end'
ws['E1'] = 'source'

# Save changes to a new file
wb.save("MICS_UNICEF_cleaned.xlsx")
