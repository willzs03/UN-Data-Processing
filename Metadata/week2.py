# import pandas
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# df = pandas.read_excel('IHSN_Metadata.xlsx')
#
# df.insert(19, "start_year", "")
#
# # for index, row in df.iterrows():
# #     year_start = str(row['study_desc.study_info.coll_dates'])[12:16]
# #     row['start_year'] = year_start
# #     print(year_start)
#
# df.to_excel('test.xlsx', index=False)

wb = load_workbook('IHSN_Metadata.xlsx')
ws = wb.active

ws.insert_cols(20)

for cell in ws['S']:
    nextCol = get_column_letter(cell.column + 1) + str(cell.row)
    toCopy = str(cell.value)[12:16]
    ws[nextCol] = toCopy

ws['T1'] = 'year_start'
wb.save('test.xlsx')


