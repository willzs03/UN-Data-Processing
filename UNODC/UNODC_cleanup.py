from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import  Font
import pandas as pd


wb = load_workbook("UNODC_SurveyData.xlsx")
ws = wb['Atlas_2021']


# Delete columns A(1), B(2), C(3), F(6), H(8)-AG(34)
ws.delete_cols(1, 3)
ws.delete_cols(3)
ws.delete_cols(4, 26)

# Move columns to corresponding places
ws.move_range("A1:A300", cols=3)
ws.move_range("C1:C300", cols=-2)
ws.move_range("B1:B300", cols=1)
ws.move_range("D1:D300", cols=-2)

wb.save("UNODC_cleaned.xlsx")

# convert all cells from int to string
df = pd.read_excel("UNODC_cleaned.xlsx", index_col=0)
df = df.applymap(str)
df.to_excel("UNODC_cleaned.xlsx")

wb = load_workbook("UNODC_cleaned.xlsx")
ws = wb['Sheet1']

# Format columns C1 (Start Year) and D1 (End Year) appropriately
for cell in ws['C']:
    currentCol = get_column_letter(cell.column) + str(cell.row)
    nextCol = get_column_letter(cell.column + 1) + str(cell.row)
    if "-" not in cell.value:
        ws[currentCol] = cell.value
        ws[nextCol] = cell.value


    # Cases with a range. e.g. 2011-2015 -> Start Year = 2011, End Year = 2015
    else:
        # Split range into two items. e.g. '2011-2015' -> ['2011', '2015']
        range = cell.value.split('-')
        ws[currentCol] = range[0]
        ws[nextCol] = range[1]

# Update 'Source' column
for cell in ws['D']:
    nextCol = get_column_letter(cell.column + 1) + str(cell.row)
    ws[nextCol] = 'UNODC'

# Undo bold font
for cell in ws['A']:
    cell.font = Font(bold=False)
ws['B1'].font = Font(bold=False)
ws['C1'].font = Font(bold=False)

# Rename column titles for consistency
ws['A1'] = 'Title'
ws['B1'] = 'Country'
ws['C1'] = 'Start Year'
ws['D1'] = 'End Year'
ws['E1'] = 'Source'

# Save changes to a new file
wb.save("UNODC_cleaned.xlsx")
