from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import  Font
import pandas as pd
from googletrans import Translator

translator = Translator()

wb = load_workbook("UNODC_cleaned.xlsx")
ws = wb['Sheet1']

out = translator.translate("Este es mi amigo", dest='en')
print(out)

for cell in ws['A']:
    text = translator.translate(cell.value, dest='en')
    # print(text.text)
    cell.value = text.text

for cell in ws['B']:
    text = translator.translate(cell.value, dest='en')
    # print(text.text)
    cell.value = text.text

wb.save("testing.xlsx")
