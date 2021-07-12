from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import  PatternFill, Border, Side, Alignment, Protection, Font

wb = load_workbook("UNWomen_SurveyData.xlsx")
ws = wb.active

# Unmerge title
ws.unmerge_cells("A1:H1")
ws.unmerge_cells("A2:H2")

# delete first three rows, not needed
ws.delete_rows(1, 3)
ws.delete_cols(4, 5)

for cell in ws['C']:
    currentCol = get_column_letter(cell.column) + str(cell.row)
    nextCol = get_column_letter(cell.column + 1) + str(cell.row)
    if "-" not in cell.value:
        ws[currentCol] = cell.value
        ws[nextCol] = cell.value


    # Cases with a range. e.g. 2011-2015 -> Start Year = 2011, End Year = 2015
    else:
        # Split range into two items. e.g. '2011-2015' -> ['2011', '2015']
        range = cell.value.split('-')
        ws[currentCol] = range[0]
        ws[nextCol] = range[1]

# Update 'Source' column
for cell in ws['D']:
    nextCol = get_column_letter(cell.column + 1) + str(cell.row)
    ws[nextCol] = 'UNWomen'

# Rename column titles for consistency
ws['A1'] = 'nation'
ws['B1'] = 'title'
ws['C1'] = 'year_start'
ws['D1'] = 'year_end'
ws['E1'] = 'source'

wb.save('UNWomen_cleaned.xlsx')